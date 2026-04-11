"""Export DiffResult to HTML and Excel for human review.

Usage::

    from pygaeb import BoQDiff, GAEBParser
    from pygaeb.diff.exports import diff_to_html, diff_to_excel

    doc_a = GAEBParser.parse("v1.X83")
    doc_b = GAEBParser.parse("v2.X83")
    result = BoQDiff.compare(doc_a, doc_b)

    diff_to_html(result, "changes.html")
    diff_to_excel(result, "changes.xlsx")
"""

from __future__ import annotations

import html
from pathlib import Path
from typing import Any

from pygaeb.diff.models import DiffResult, ItemModified


def diff_to_html(result: DiffResult, path: str | Path) -> None:
    """Render a DiffResult as a standalone HTML report.

    The output is a single self-contained HTML file with inline CSS,
    safe to email or open in any browser.

    Args:
        result: The :class:`DiffResult` from :meth:`BoQDiff.compare`.
        path: Output file path.
    """
    out = Path(path)
    out.write_text(_render_html(result), encoding="utf-8")


def diff_to_excel(result: DiffResult, path: str | Path) -> None:
    """Export a DiffResult to a color-coded Excel workbook.

    Three sheets: Summary, Items (added/removed/modified), Structure.

    Args:
        result: The :class:`DiffResult` from :meth:`BoQDiff.compare`.
        path: Output file path (.xlsx).

    Raises:
        ImportError: If ``openpyxl`` is not installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as e:
        raise ImportError(
            "Diff Excel export requires openpyxl. "
            "Install with: pip install pyGAEB[excel]"
        ) from e

    wb = Workbook()

    # Sheet 1: Summary
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Summary")
    else:
        ws.title = "Summary"
    bold = Font(bold=True)
    ws["A1"] = "Diff Summary"
    ws["A1"].font = Font(bold=True, size=14)
    rows: list[tuple[str, Any]] = [
        ("", ""),
        ("Total changes", result.summary.total_changes),
        ("Items added", result.summary.items_added),
        ("Items removed", result.summary.items_removed),
        ("Items modified", result.summary.items_modified),
        ("Items unchanged", result.summary.items_unchanged),
        ("Match ratio", f"{result.summary.match_ratio:.1%}"),
        ("Max significance", result.summary.max_significance.value),
    ]
    if result.summary.financial_impact is not None:
        # Pass Decimal directly — openpyxl preserves Decimal precision when
        # writing cells. Converting to float() would silently lose precision
        # on large or 3-decimal totals.
        rows.append(("Financial impact", result.summary.financial_impact))
    for i, (label, value) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label).font = bold
        ws.cell(row=i, column=2, value=value)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28

    # Sheet 2: Items
    items_ws = wb.create_sheet("Items")
    headers = ["Status", "OZ", "Description", "Significance", "Field changes"]
    for col, h in enumerate(headers, start=1):
        cell = items_ws.cell(row=1, column=col, value=h)
        cell.font = bold

    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    yellow = PatternFill("solid", fgColor="FFEB9C")

    row_idx = 2
    for added in result.items.added:
        for col, val in enumerate(
            ["ADDED", added.oz, added.short_text, "—", ""], start=1
        ):
            cell = items_ws.cell(row=row_idx, column=col, value=val)
            cell.fill = green
        row_idx += 1
    for removed in result.items.removed:
        for col, val in enumerate(
            ["REMOVED", removed.oz, removed.short_text, "—", ""], start=1
        ):
            cell = items_ws.cell(row=row_idx, column=col, value=val)
            cell.fill = red
        row_idx += 1
    for modified in result.items.modified:
        sig = modified.max_significance.value
        change_summary = "; ".join(
            f"{c.field}: {c.old_value} -> {c.new_value}"
            for c in modified.changes
        )
        for col, val in enumerate(
            ["MODIFIED", modified.oz, modified.short_text_b or modified.short_text_a,
             sig, change_summary],
            start=1,
        ):
            cell = items_ws.cell(row=row_idx, column=col, value=val)
            cell.fill = yellow
        row_idx += 1

    items_ws.column_dimensions["A"].width = 12
    items_ws.column_dimensions["B"].width = 15
    items_ws.column_dimensions["C"].width = 50
    items_ws.column_dimensions["D"].width = 12
    items_ws.column_dimensions["E"].width = 80
    items_ws.freeze_panes = "A2"

    # Sheet 3: Structure
    struct_ws = wb.create_sheet("Structure")
    struct_ws["A1"] = "Section"
    struct_ws["B1"] = "Change"
    struct_ws["C1"] = "Details"
    for cell_addr in ("A1", "B1", "C1"):
        struct_ws[cell_addr].font = bold
    sr = 2
    for s in result.structure.sections_added:
        struct_ws.cell(row=sr, column=1, value=s.rno)
        struct_ws.cell(row=sr, column=2, value="ADDED").fill = green
        struct_ws.cell(row=sr, column=3, value=s.label)
        sr += 1
    for s in result.structure.sections_removed:
        struct_ws.cell(row=sr, column=1, value=s.rno)
        struct_ws.cell(row=sr, column=2, value="REMOVED").fill = red
        struct_ws.cell(row=sr, column=3, value=s.label)
        sr += 1
    for sr_change in result.structure.sections_renamed:
        struct_ws.cell(row=sr, column=1, value=sr_change.rno)
        struct_ws.cell(row=sr, column=2, value="RENAMED").fill = yellow
        struct_ws.cell(
            row=sr, column=3,
            value=f"{sr_change.old_label} -> {sr_change.new_label}",
        )
        sr += 1
    struct_ws.column_dimensions["A"].width = 12
    struct_ws.column_dimensions["B"].width = 12
    struct_ws.column_dimensions["C"].width = 60

    wb.save(str(path))


_EMPTY_ROW = (
    '<tr><td colspan="5" style="text-align:center;color:#888;padding:2em">'
    "No item-level changes detected.</td></tr>"
)


def _render_html(result: DiffResult) -> str:
    """Render a DiffResult to a self-contained HTML string."""
    s = result.summary
    rows: list[str] = []

    for added in result.items.added:
        rows.append(
            f'<tr class="added"><td>+ ADDED</td><td>{html.escape(added.oz)}</td>'
            f'<td>{html.escape(added.short_text or "")}</td><td>—</td><td></td></tr>'
        )
    for removed in result.items.removed:
        rows.append(
            f'<tr class="removed"><td>- REMOVED</td><td>{html.escape(removed.oz)}</td>'
            f'<td>{html.escape(removed.short_text or "")}</td><td>—</td><td></td></tr>'
        )
    for modified in result.items.modified:
        rows.append(_render_modified_row(modified))

    fin = ""
    if s.financial_impact is not None:
        sign = "+" if s.financial_impact >= 0 else ""
        fin = (
            f'<div class="metric"><span>Financial impact</span>'
            f'<strong>{sign}{s.financial_impact:,.2f}</strong></div>'
        )

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>pyGAEB Diff Report</title>
<style>
  body {{ font-family: -apple-system, sans-serif; margin: 2em; color: #222; }}
  h1 {{ border-bottom: 2px solid #333; padding-bottom: 0.3em; }}
  .summary {{ display: flex; gap: 1.5em; flex-wrap: wrap; margin: 1em 0 2em; }}
  .metric {{
    background: #f4f4f4; padding: 0.8em 1.2em; border-radius: 6px;
    display: flex; flex-direction: column;
  }}
  .metric span {{ font-size: 0.85em; color: #666; }}
  .metric strong {{ font-size: 1.4em; margin-top: 0.2em; }}
  table {{ border-collapse: collapse; width: 100%; font-size: 0.92em; }}
  th, td {{ padding: 0.5em 0.8em; border-bottom: 1px solid #ddd; text-align: left; }}
  th {{ background: #333; color: white; }}
  tr.added td {{ background: #e6ffec; }}
  tr.removed td {{ background: #ffeef0; }}
  tr.modified td {{ background: #fff8c5; }}
  .field-change {{ font-family: monospace; font-size: 0.85em; color: #555; }}
  .significance-critical {{ color: #d00; font-weight: bold; }}
  .significance-high {{ color: #e50; font-weight: bold; }}
  .significance-medium {{ color: #c80; }}
  .significance-low {{ color: #888; }}
</style>
</head>
<body>
<h1>pyGAEB Diff Report</h1>
<div class="summary">
  <div class="metric"><span>Total changes</span><strong>{s.total_changes}</strong></div>
  <div class="metric"><span>Added</span><strong>{s.items_added}</strong></div>
  <div class="metric"><span>Removed</span><strong>{s.items_removed}</strong></div>
  <div class="metric"><span>Modified</span><strong>{s.items_modified}</strong></div>
  <div class="metric"><span>Unchanged</span><strong>{s.items_unchanged}</strong></div>
  <div class="metric"><span>Match ratio</span><strong>{s.match_ratio:.1%}</strong></div>
  <div class="metric"><span>Max significance</span><strong>{s.max_significance.value}</strong></div>
  {fin}
</div>
<table>
<thead>
<tr><th>Status</th><th>OZ</th><th>Description</th><th>Significance</th><th>Changes</th></tr>
</thead>
<tbody>
{"".join(rows) if rows else _EMPTY_ROW}
</tbody>
</table>
</body>
</html>
"""


def _render_modified_row(modified: ItemModified) -> str:
    """Render a single modified item row."""
    sig = modified.max_significance
    sig_class = f"significance-{sig.value}"
    changes_html = "<br>".join(
        f'<span class="field-change">{html.escape(c.field)}: '
        f'{html.escape(str(c.old_value))} → {html.escape(str(c.new_value))}</span>'
        for c in modified.changes
    )
    short_text = html.escape(modified.short_text_b or modified.short_text_a or "")
    return (
        f'<tr class="modified"><td>~ MODIFIED</td>'
        f'<td>{html.escape(modified.oz)}</td>'
        f'<td>{short_text}</td>'
        f'<td class="{sig_class}">{sig.value}</td>'
        f'<td>{changes_html}</td></tr>'
    )


__all__ = ["diff_to_excel", "diff_to_html"]
