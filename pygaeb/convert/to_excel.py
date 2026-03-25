"""Export GAEBDocument to Excel (.xlsx) with structure-aware layout.

Supports all four document kinds (procurement, trade, cost, quantity) and
two modes: ``structured`` (single hierarchy-aware sheet) or ``full``
(multi-sheet workbook with BoQ, Items, Summary, and Info sheets).

Requires ``openpyxl``: ``pip install pyGAEB[excel]``.
"""

from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any

from pygaeb.models.document import GAEBDocument
from pygaeb.models.enums import DocumentKind

_CURRENCY_FMT = '#,##0.00'
_QTY_FMT = '#,##0.###'
_MAX_COL_WIDTH = 60
_MIN_COL_WIDTH = 8


@dataclass(frozen=True)
class ColumnDef:
    """Definition of an Excel column."""

    key: str
    header: str
    width: int = 14
    number_format: str | None = None


_PROCUREMENT_COLS = [
    ColumnDef("oz", "OZ", 14),
    ColumnDef("short_text", "Description", 40),
    ColumnDef("qty", "Qty", 12, _QTY_FMT),
    ColumnDef("unit", "Unit", 8),
    ColumnDef("unit_price", "Unit Price", 14, _CURRENCY_FMT),
    ColumnDef("total_price", "Total Price", 16, _CURRENCY_FMT),
    ColumnDef("computed_total", "Computed Total", 16, _CURRENCY_FMT),
    ColumnDef("item_type", "Item Type", 14),
    ColumnDef("hierarchy_path", "Hierarchy", 30),
]

_TRADE_COLS = [
    ColumnDef("item_id", "Item ID", 14),
    ColumnDef("art_no", "Article No.", 16),
    ColumnDef("short_text", "Description", 40),
    ColumnDef("qty", "Qty", 12, _QTY_FMT),
    ColumnDef("unit", "Unit", 8),
    ColumnDef("offer_price", "Offer Price", 14, _CURRENCY_FMT),
    ColumnDef("net_price", "Net Price", 14, _CURRENCY_FMT),
    ColumnDef("is_service", "Service", 10),
]

_COST_COLS = [
    ColumnDef("ele_no", "Element No.", 14),
    ColumnDef("short_text", "Description", 40),
    ColumnDef("qty", "Qty", 12, _QTY_FMT),
    ColumnDef("unit", "Unit", 8),
    ColumnDef("unit_price", "Unit Price", 14, _CURRENCY_FMT),
    ColumnDef("item_total", "Total", 16, _CURRENCY_FMT),
    ColumnDef("markup", "Markup", 12, _QTY_FMT),
    ColumnDef("cat_id", "Category ID", 14),
]

_QTY_COLS = [
    ColumnDef("oz", "OZ", 14),
    ColumnDef("rno_part", "RNo Part", 12),
    ColumnDef("qty", "Qty", 12, _QTY_FMT),
    ColumnDef("determ_count", "Determinations", 16),
]

_OPTIONAL_COLS = {
    "long_text": ColumnDef("long_text_plain", "Long Text", 50),
    "classification_trade": ColumnDef("classification_trade", "Trade", 16),
    "classification_element": ColumnDef("classification_element_type", "Element Type", 20),
    "classification_confidence": ColumnDef("classification_confidence", "Confidence", 12, '0.00'),
    "bim_guid": ColumnDef("bim_guid", "BIM GUID", 36),
}

_COLUMNS_BY_KIND = {
    DocumentKind.PROCUREMENT: _PROCUREMENT_COLS,
    DocumentKind.TRADE: _TRADE_COLS,
    DocumentKind.COST: _COST_COLS,
    DocumentKind.QUANTITY: _QTY_COLS,
}


def _ensure_openpyxl() -> Any:
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        raise ImportError(
            "Excel export requires openpyxl. Install it with: pip install pyGAEB[excel]"
        ) from None


def _get_columns(
    doc: GAEBDocument,
    include_long_text: bool,
    include_classification: bool,
    include_bim_guid: bool,
) -> list[ColumnDef]:
    cols = list(_COLUMNS_BY_KIND[doc.document_kind])
    if include_long_text:
        cols.append(_OPTIONAL_COLS["long_text"])
    if include_classification:
        cols.append(_OPTIONAL_COLS["classification_trade"])
        cols.append(_OPTIONAL_COLS["classification_element"])
        cols.append(_OPTIONAL_COLS["classification_confidence"])
    if include_bim_guid and doc.document_kind == DocumentKind.PROCUREMENT:
        cols.append(_OPTIONAL_COLS["bim_guid"])
    return cols


def _get_item_value(item: Any, col: ColumnDef) -> Any:
    """Extract a value from an item for a given column definition."""
    key = col.key
    if key == "computed_total":
        return getattr(item, "computed_total", None)
    if key == "hierarchy_path":
        path = getattr(item, "hierarchy_path", None)
        return " > ".join(path) if path else ""
    if key == "item_type":
        it = getattr(item, "item_type", None)
        return it.value if it is not None else ""
    if key == "determ_count":
        determ = getattr(item, "determ_items", None)
        return len(determ) if determ else 0
    if key == "long_text_plain":
        return getattr(item, "long_text_plain", "")
    if key.startswith("classification_"):
        cls_obj = getattr(item, "classification", None)
        if cls_obj is None:
            return ""
        suffix = key.replace("classification_", "")
        return getattr(cls_obj, suffix, "")
    if key == "is_service":
        return "Yes" if getattr(item, "is_service", False) else ""

    val = getattr(item, key, None)
    if isinstance(val, Decimal):
        return float(val)
    return val if val is not None else ""


def to_excel(
    doc: GAEBDocument,
    path: str | Path,
    *,
    mode: str = "structured",
    include_long_text: bool = False,
    include_classification: bool = False,
    include_bim_guid: bool = False,
) -> None:
    """Export a GAEBDocument to an Excel workbook.

    Args:
        doc: The document to export.
        path: Output file path (.xlsx).
        mode: ``"structured"`` for a single hierarchy-aware sheet, or
              ``"full"`` for a multi-sheet workbook (BoQ + Items + Summary + Info).
        include_long_text: Add a "Long Text" column.
        include_classification: Add classification columns (trade, element_type, confidence).
        include_bim_guid: Add a "BIM GUID" column (procurement only).
    """
    openpyxl = _ensure_openpyxl()
    if mode not in ("structured", "full"):
        raise ValueError(f"Invalid mode {mode!r}. Use 'structured' or 'full'.")

    wb = openpyxl.Workbook()
    columns = _get_columns(doc, include_long_text, include_classification, include_bim_guid)

    ws_boq = wb.active
    ws_boq.title = "BoQ"
    _write_structured_sheet(ws_boq, doc, columns)
    _apply_formatting(ws_boq, columns)

    if mode == "full":
        ws_items = wb.create_sheet("Items")
        _write_flat_items(ws_items, doc, columns)
        _apply_formatting(ws_items, columns)

        ws_summary = wb.create_sheet("Summary")
        _write_summary_sheet(ws_summary, doc)
        _apply_formatting_auto(ws_summary)

        ws_info = wb.create_sheet("Info")
        _write_info_sheet(ws_info, doc)
        _apply_formatting_auto(ws_info)

    wb.save(Path(path))


# ---------------------------------------------------------------------------
# Structured sheet writers (dispatched by DocumentKind)
# ---------------------------------------------------------------------------

def _write_structured_sheet(ws: Any, doc: GAEBDocument, columns: list[ColumnDef]) -> None:
    kind = doc.document_kind
    writers = {
        DocumentKind.PROCUREMENT: _write_procurement_structured,
        DocumentKind.TRADE: _write_trade_structured,
        DocumentKind.COST: _write_cost_structured,
        DocumentKind.QUANTITY: _write_quantity_structured,
    }
    writers[kind](ws, doc, columns)


def _write_header_row(ws: Any, columns: list[ColumnDef], row: int = 1) -> int:
    from openpyxl.styles import Font

    for col_idx, col_def in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx, value=col_def.header)
        cell.font = Font(bold=True)
    return row + 1


def _write_item_row(
    ws: Any, row: int, item: Any, columns: list[ColumnDef]
) -> int:
    for col_idx, col_def in enumerate(columns, 1):
        val = _get_item_value(item, col_def)
        cell = ws.cell(row=row, column=col_idx, value=val)
        if col_def.number_format and isinstance(val, (int, float)):
            cell.number_format = col_def.number_format
    return row + 1


def _write_label_row(
    ws: Any, row: int, label: str, columns: list[ColumnDef], *, bold: bool = True
) -> int:
    from openpyxl.styles import Font

    cell = ws.cell(row=row, column=1, value=label)
    cell.font = Font(bold=bold)
    return row + 1


def _write_subtotal_row(
    ws: Any, row: int, label: str, value: Decimal | float, columns: list[ColumnDef]
) -> int:
    from openpyxl.styles import Font

    total_col = _find_total_column(columns)
    cell_label = ws.cell(row=row, column=max(1, total_col - 1), value=label)
    cell_label.font = Font(bold=True)
    val = float(value) if isinstance(value, Decimal) else value
    cell_val = ws.cell(row=row, column=total_col, value=val)
    cell_val.font = Font(bold=True)
    cell_val.number_format = _CURRENCY_FMT
    return row + 1


def _find_total_column(columns: list[ColumnDef]) -> int:
    for i, c in enumerate(columns, 1):
        if c.key in ("total_price", "net_price", "item_total", "qty"):
            return i
    return len(columns)


# ---------------------------------------------------------------------------
# Procurement structured writer
# ---------------------------------------------------------------------------

def _write_procurement_structured(
    ws: Any, doc: GAEBDocument, columns: list[ColumnDef]
) -> None:
    row = _write_header_row(ws, columns)

    for lot in doc.award.boq.lots:
        lot_label = f"Lot {lot.rno}"
        if lot.label:
            lot_label += f" — {lot.label}"
        row = _write_label_row(ws, row, lot_label, columns)

        for ctgy in lot.body.categories:
            row = _write_procurement_category(ws, row, ctgy, columns, depth=1)

        lot_total = float(lot.subtotal)
        if lot_total:
            row = _write_subtotal_row(ws, row, "Lot Subtotal", lot_total, columns)

    grand = float(doc.grand_total)
    if grand:
        row = _write_subtotal_row(ws, row, "GRAND TOTAL", grand, columns)


def _write_procurement_category(
    ws: Any, row: int, ctgy: Any, columns: list[ColumnDef], depth: int
) -> int:
    indent = "  " * depth
    label = f"{indent}{ctgy.rno}"
    if ctgy.label:
        label += f" — {ctgy.label}"
    row = _write_label_row(ws, row, label, columns)

    for item in ctgy.items:
        row = _write_item_row(ws, row, item, columns)

    for sub in ctgy.subcategories:
        row = _write_procurement_category(ws, row, sub, columns, depth + 1)

    subtotal = float(ctgy.subtotal)
    if subtotal and (ctgy.items or ctgy.subcategories):
        row = _write_subtotal_row(
            ws, row, f"{indent}Subtotal {ctgy.rno}", subtotal, columns
        )

    return row


# ---------------------------------------------------------------------------
# Trade structured writer
# ---------------------------------------------------------------------------

def _write_trade_structured(
    ws: Any, doc: GAEBDocument, columns: list[ColumnDef]
) -> None:
    row = _write_header_row(ws, columns)

    if doc.order is None:
        return

    for item in doc.order.items:
        row = _write_item_row(ws, row, item, columns)

    grand = float(doc.order.grand_total)
    if grand:
        row = _write_subtotal_row(ws, row, "TOTAL", grand, columns)


# ---------------------------------------------------------------------------
# Cost structured writer
# ---------------------------------------------------------------------------

def _write_cost_structured(
    ws: Any, doc: GAEBDocument, columns: list[ColumnDef]
) -> None:
    row = _write_header_row(ws, columns)

    if doc.elemental_costing is None:
        return

    ec = doc.elemental_costing
    for ctgy in ec.body.categories:
        row = _write_cost_category(ws, row, ctgy, columns, depth=0)

    for ce in ec.body.cost_elements:
        row = _write_cost_element(ws, row, ce, columns, depth=0)

    grand = float(ec.grand_total)
    if grand:
        row = _write_subtotal_row(ws, row, "GRAND TOTAL", grand, columns)


def _write_cost_category(
    ws: Any, row: int, ctgy: Any, columns: list[ColumnDef], depth: int
) -> int:
    indent = "  " * depth
    label = f"{indent}{ctgy.ele_no}"
    if ctgy.description:
        label += f" — {ctgy.description}"
    row = _write_label_row(ws, row, label, columns)

    if ctgy.body is not None:
        for sub in ctgy.body.categories:
            row = _write_cost_category(ws, row, sub, columns, depth + 1)
        for ce in ctgy.body.cost_elements:
            row = _write_cost_element(ws, row, ce, columns, depth + 1)

    return row


def _write_cost_element(
    ws: Any, row: int, ce: Any, columns: list[ColumnDef], depth: int
) -> int:
    row = _write_item_row(ws, row, ce, columns)
    for child in ce.children:
        row = _write_cost_element(ws, row, child, columns, depth + 1)
    return row


# ---------------------------------------------------------------------------
# Quantity structured writer
# ---------------------------------------------------------------------------

def _write_quantity_structured(
    ws: Any, doc: GAEBDocument, columns: list[ColumnDef]
) -> None:
    row = _write_header_row(ws, columns)

    if doc.qty_determination is None:
        return

    qd = doc.qty_determination
    for ctgy in qd.boq.body.categories:
        row = _write_qty_category(ws, row, ctgy, columns, depth=0)


def _write_qty_category(
    ws: Any, row: int, ctgy: Any, columns: list[ColumnDef], depth: int
) -> int:
    indent = "  " * depth
    label = f"{indent}Category {ctgy.rno}"
    row = _write_label_row(ws, row, label, columns)

    for item in ctgy.items:
        row = _write_item_row(ws, row, item, columns)

    for sub in ctgy.subcategories:
        row = _write_qty_category(ws, row, sub, columns, depth + 1)

    return row


# ---------------------------------------------------------------------------
# Flat items sheet (shared)
# ---------------------------------------------------------------------------

def _write_flat_items(
    ws: Any, doc: GAEBDocument, columns: list[ColumnDef]
) -> None:
    row = _write_header_row(ws, columns)
    for item in doc.iter_items():
        row = _write_item_row(ws, row, item, columns)


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

def _write_summary_sheet(ws: Any, doc: GAEBDocument) -> None:
    from openpyxl.styles import Font

    headers = ["Metric", "Value"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)

    row = 2

    if doc.is_procurement:
        for lot in doc.award.boq.lots:
            lot_label = f"Lot {lot.rno}"
            if lot.label:
                lot_label += f" — {lot.label}"
            ws.cell(row=row, column=1, value=lot_label)
            ws.cell(row=row, column=2)
            row += 1

            item_count = sum(1 for _ in lot.iter_items())
            ws.cell(row=row, column=1, value="  Items")
            ws.cell(row=row, column=2, value=item_count)
            row += 1

            subtotal = float(lot.subtotal)
            ws.cell(row=row, column=1, value="  Subtotal")
            cell = ws.cell(row=row, column=2, value=subtotal)
            cell.number_format = _CURRENCY_FMT
            row += 1

        row += 1

    ws.cell(row=row, column=1, value="Total Items")
    ws.cell(row=row, column=2, value=doc.item_count)
    row += 1

    cell_label = ws.cell(row=row, column=1, value="Grand Total")
    cell_label.font = Font(bold=True)
    cell_val = ws.cell(row=row, column=2, value=float(doc.grand_total))
    cell_val.font = Font(bold=True)
    cell_val.number_format = _CURRENCY_FMT


# ---------------------------------------------------------------------------
# Info sheet
# ---------------------------------------------------------------------------

def _write_info_sheet(ws: Any, doc: GAEBDocument) -> None:
    from openpyxl.styles import Font

    headers = ["Property", "Value"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)

    info_rows: list[tuple[str, Any]] = [
        ("Document Kind", doc.document_kind.value),
        ("Exchange Phase", doc.exchange_phase.value),
        ("Source Version", doc.source_version.value),
        ("Source File", doc.source_file or ""),
    ]

    if doc.gaeb_info.prog_system:
        info_rows.append(("Program System", doc.gaeb_info.prog_system))
    if doc.gaeb_info.prog_system_version:
        info_rows.append(("Program Version", doc.gaeb_info.prog_system_version))

    if doc.is_procurement:
        award = doc.award
        if award.project_name:
            info_rows.append(("Project Name", award.project_name))
        if award.project_no:
            info_rows.append(("Project No.", award.project_no))
        if award.currency:
            info_rows.append(("Currency", award.currency))
        if award.client:
            info_rows.append(("Client", award.client))

    if doc.is_trade and doc.order is not None:
        oi = doc.order.order_info
        if oi is not None:
            if oi.order_no:
                info_rows.append(("Order No.", oi.order_no))
            if oi.currency:
                info_rows.append(("Currency", oi.currency))

    if doc.is_cost and doc.elemental_costing is not None:
        ec_info = doc.elemental_costing.ec_info
        if ec_info.name:
            info_rows.append(("Cost Estimation Name", ec_info.name))
        if ec_info.currency:
            info_rows.append(("Currency", ec_info.currency))

    for row_idx, (prop, val) in enumerate(info_rows, 2):
        ws.cell(row=row_idx, column=1, value=prop)
        ws.cell(row=row_idx, column=2, value=str(val) if val else "")


# ---------------------------------------------------------------------------
# Formatting
# ---------------------------------------------------------------------------

def _apply_formatting(ws: Any, columns: list[ColumnDef]) -> None:
    """Apply minimal formatting: column widths and frozen header."""
    for col_idx, col_def in enumerate(columns, 1):
        max_len = col_def.width
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, min(len(str(cell.value)), _MAX_COL_WIDTH))
        col_letter = _col_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(max_len + 2, _MIN_COL_WIDTH)

    ws.freeze_panes = "A2"


def _apply_formatting_auto(ws: Any) -> None:
    """Auto-width for simple sheets (Summary, Info)."""
    for col in ws.columns:
        max_len = _MIN_COL_WIDTH
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), _MAX_COL_WIDTH))
        ws.column_dimensions[col_letter].width = max_len + 2

    ws.freeze_panes = "A2"


def _col_letter(col_idx: int) -> str:
    from openpyxl.utils import get_column_letter
    return str(get_column_letter(col_idx))
