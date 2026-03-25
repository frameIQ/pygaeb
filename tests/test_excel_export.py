"""Tests for the Excel export feature."""

from __future__ import annotations  # noqa: I001

from decimal import Decimal
from pathlib import Path

import pytest

from pygaeb.convert.to_excel import to_excel
from pygaeb.models.boq import BoQ, BoQBody, BoQCtgy, BoQInfo, Lot, Totals
from pygaeb.models.cost import (
    CostElement,
    ECBody,
    ECCtgy,
    ECInfo,
    ElementalCosting,
)
from pygaeb.models.document import AwardInfo, GAEBDocument, GAEBInfo
from pygaeb.models.enums import (
    ExchangePhase,
    ItemType,
    SourceVersion,
)
from pygaeb.models.item import ClassificationResult, Item
from pygaeb.models.order import OrderInfo, OrderItem, TradeOrder
from pygaeb.models.quantity import (
    QtyBoQ,
    QtyBoQBody,
    QtyBoQCtgy,
    QtyDetermination,
    QtyItem,
)


# ── Fixtures ────────────────────────────────────────────────────────────

def _make_procurement_doc() -> GAEBDocument:
    items_a = [
        Item(oz="01.0010", short_text="Foundation", qty=Decimal("120"),
             unit="m3", unit_price=Decimal("85"), total_price=Decimal("10200")),
        Item(oz="01.0020", short_text="Columns", qty=Decimal("40"),
             unit="m3", unit_price=Decimal("95"), total_price=Decimal("3800")),
    ]
    items_b = [
        Item(oz="02.0010", short_text="Exterior walls", qty=Decimal("600"),
             unit="m2", unit_price=Decimal("68"), total_price=Decimal("40800")),
    ]
    ctgy_a = BoQCtgy(rno="01", label="Concrete", items=items_a)
    ctgy_b = BoQCtgy(rno="02", label="Masonry", items=items_b)
    lot = Lot(rno="1", label="Structural Work",
              body=BoQBody(categories=[ctgy_a, ctgy_b]),
              totals=Totals(total=Decimal("54800")))
    boq = BoQ(boq_info=BoQInfo(), lots=[lot])
    award = AwardInfo(
        boq=boq, project_name="Test Project", project_no="PRJ-001", currency="EUR"
    )
    return GAEBDocument(
        source_version=SourceVersion.DA_XML_33,
        exchange_phase=ExchangePhase.X83,
        gaeb_info=GAEBInfo(prog_system="pyGAEB-Test"),
        award=award,
    )


def _make_trade_doc() -> GAEBDocument:
    items = [
        OrderItem(
            item_id="1", art_no="W-100", short_text="Window frame",
            qty=Decimal("10"), unit="Stk",
            offer_price=Decimal("250"), net_price=Decimal("230"),
        ),
        OrderItem(
            item_id="2", art_no="D-200", short_text="Door panel",
            qty=Decimal("5"), unit="Stk",
            offer_price=Decimal("400"), net_price=Decimal("380"),
            is_service=True,
        ),
    ]
    order = TradeOrder(
        dp="X96",
        order_info=OrderInfo(order_no="ORD-001", currency="EUR"),
        items=items,
    )
    return GAEBDocument(
        source_version=SourceVersion.DA_XML_33,
        exchange_phase=ExchangePhase.X96,
        gaeb_info=GAEBInfo(),
        order=order,
    )


def _make_cost_doc() -> GAEBDocument:
    ce1 = CostElement(
        ele_no="310", short_text="Walls", qty=Decimal("100"),
        unit="m2", unit_price=Decimal("85"), item_total=Decimal("8500"),
    )
    ce2 = CostElement(
        ele_no="320", short_text="Floors", qty=Decimal("200"),
        unit="m2", unit_price=Decimal("65"), item_total=Decimal("13000"),
    )
    ctgy = ECCtgy(
        ele_no="300", description="Building Construction",
        body=ECBody(cost_elements=[ce1, ce2]),
    )
    ec = ElementalCosting(
        dp="X50",
        ec_info=ECInfo(name="Office Building", currency="EUR"),
        body=ECBody(categories=[ctgy]),
    )
    return GAEBDocument(
        source_version=SourceVersion.DA_XML_33,
        exchange_phase=ExchangePhase.X50,
        gaeb_info=GAEBInfo(),
        elemental_costing=ec,
    )


def _make_qty_doc() -> GAEBDocument:
    qi1 = QtyItem(oz="01.0010", rno_part="01", qty=Decimal("120"))
    qi2 = QtyItem(oz="01.0020", rno_part="01", qty=Decimal("80"))
    ctgy = QtyBoQCtgy(rno="01", items=[qi1, qi2])
    qd = QtyDetermination(
        dp="X31",
        boq=QtyBoQ(body=QtyBoQBody(categories=[ctgy])),
    )
    return GAEBDocument(
        source_version=SourceVersion.DA_XML_33,
        exchange_phase=ExchangePhase.X31,
        gaeb_info=GAEBInfo(),
        qty_determination=qd,
    )


# ── Helper ──────────────────────────────────────────────────────────────

def _load_workbook(path: Path):
    import openpyxl
    return openpyxl.load_workbook(path)


# ── Procurement Tests ───────────────────────────────────────────────────

class TestProcurementExport:
    def test_structured_creates_file(self, tmp_path: Path):
        doc = _make_procurement_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out, mode="structured")
        assert out.exists()
        assert out.stat().st_size > 0

    def test_structured_has_single_sheet(self, tmp_path: Path):
        doc = _make_procurement_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out, mode="structured")
        wb = _load_workbook(out)
        assert len(wb.sheetnames) == 1
        assert wb.sheetnames[0] == "BoQ"

    def test_structured_contains_items(self, tmp_path: Path):
        doc = _make_procurement_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out, mode="structured")
        wb = _load_workbook(out)
        ws = wb["BoQ"]
        values = [row[0] for row in ws.iter_rows(values_only=True)]
        assert "OZ" in values
        assert any("Lot 1" in str(v) for v in values if v)

    def test_structured_has_grand_total(self, tmp_path: Path):
        doc = _make_procurement_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out, mode="structured")
        wb = _load_workbook(out)
        ws = wb["BoQ"]
        all_vals = []
        for row in ws.iter_rows(values_only=True):
            all_vals.extend(row)
        assert any("GRAND TOTAL" in str(v) for v in all_vals if v)

    def test_full_mode_has_four_sheets(self, tmp_path: Path):
        doc = _make_procurement_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out, mode="full")
        wb = _load_workbook(out)
        assert set(wb.sheetnames) == {"BoQ", "Items", "Summary", "Info"}

    def test_full_items_sheet_has_all_items(self, tmp_path: Path):
        doc = _make_procurement_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out, mode="full")
        wb = _load_workbook(out)
        ws = wb["Items"]
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 3

    def test_full_info_sheet_has_metadata(self, tmp_path: Path):
        doc = _make_procurement_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out, mode="full")
        wb = _load_workbook(out)
        ws = wb["Info"]
        all_vals = []
        for row in ws.iter_rows(values_only=True):
            all_vals.extend(row)
        assert "Test Project" in all_vals
        assert "PRJ-001" in all_vals
        assert "EUR" in all_vals

    def test_full_summary_sheet(self, tmp_path: Path):
        doc = _make_procurement_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out, mode="full")
        wb = _load_workbook(out)
        ws = wb["Summary"]
        all_vals = []
        for row in ws.iter_rows(values_only=True):
            all_vals.extend(row)
        assert "Grand Total" in all_vals
        assert 3 in all_vals


# ── Trade Tests ─────────────────────────────────────────────────────────

class TestTradeExport:
    def test_structured_creates_file(self, tmp_path: Path):
        doc = _make_trade_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out, mode="structured")
        assert out.exists()

    def test_structured_has_trade_columns(self, tmp_path: Path):
        doc = _make_trade_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out, mode="structured")
        wb = _load_workbook(out)
        ws = wb["BoQ"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "Item ID" in headers
        assert "Article No." in headers
        assert "Net Price" in headers

    def test_service_flag(self, tmp_path: Path):
        doc = _make_trade_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out, mode="structured")
        wb = _load_workbook(out)
        ws = wb["BoQ"]
        all_vals = []
        for row in ws.iter_rows(values_only=True):
            all_vals.extend(row)
        assert "Yes" in all_vals

    def test_full_mode(self, tmp_path: Path):
        doc = _make_trade_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out, mode="full")
        wb = _load_workbook(out)
        assert "Items" in wb.sheetnames
        ws = wb["Items"]
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 2


# ── Cost Tests ──────────────────────────────────────────────────────────

class TestCostExport:
    def test_structured_creates_file(self, tmp_path: Path):
        doc = _make_cost_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out, mode="structured")
        assert out.exists()

    def test_structured_has_cost_columns(self, tmp_path: Path):
        doc = _make_cost_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out, mode="structured")
        wb = _load_workbook(out)
        ws = wb["BoQ"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "Element No." in headers
        assert "Total" in headers
        assert "Markup" in headers

    def test_structured_has_category_header(self, tmp_path: Path):
        doc = _make_cost_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out, mode="structured")
        wb = _load_workbook(out)
        ws = wb["BoQ"]
        all_vals = [row[0] for row in ws.iter_rows(values_only=True)]
        assert any("Building Construction" in str(v) for v in all_vals if v)

    def test_full_mode(self, tmp_path: Path):
        doc = _make_cost_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out, mode="full")
        wb = _load_workbook(out)
        assert set(wb.sheetnames) == {"BoQ", "Items", "Summary", "Info"}


# ── Quantity Tests ──────────────────────────────────────────────────────

class TestQuantityExport:
    def test_structured_creates_file(self, tmp_path: Path):
        doc = _make_qty_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out, mode="structured")
        assert out.exists()

    def test_structured_has_qty_columns(self, tmp_path: Path):
        doc = _make_qty_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out, mode="structured")
        wb = _load_workbook(out)
        ws = wb["BoQ"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "OZ" in headers
        assert "Determinations" in headers

    def test_full_mode(self, tmp_path: Path):
        doc = _make_qty_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out, mode="full")
        wb = _load_workbook(out)
        assert "Items" in wb.sheetnames


# ── Column Flags ────────────────────────────────────────────────────────

class TestColumnFlags:
    def test_include_long_text(self, tmp_path: Path):
        doc = _make_procurement_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out, include_long_text=True)
        wb = _load_workbook(out)
        ws = wb["BoQ"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "Long Text" in headers

    def test_include_classification(self, tmp_path: Path):
        doc = _make_procurement_doc()
        for item in doc.award.boq.iter_items():
            item.classification = ClassificationResult(
                trade="Masonry", element_type="Wall", confidence=0.95
            )
        out = tmp_path / "test.xlsx"
        to_excel(doc, out, include_classification=True)
        wb = _load_workbook(out)
        ws = wb["BoQ"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "Trade" in headers
        assert "Element Type" in headers
        assert "Confidence" in headers

    def test_include_bim_guid(self, tmp_path: Path):
        doc = _make_procurement_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out, include_bim_guid=True)
        wb = _load_workbook(out)
        ws = wb["BoQ"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "BIM GUID" in headers

    def test_bim_guid_not_added_for_trade(self, tmp_path: Path):
        doc = _make_trade_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out, include_bim_guid=True)
        wb = _load_workbook(out)
        ws = wb["BoQ"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assert "BIM GUID" not in headers


# ── Formatting Tests ────────────────────────────────────────────────────

class TestFormatting:
    def test_header_row_is_bold(self, tmp_path: Path):
        doc = _make_procurement_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out)
        wb = _load_workbook(out)
        ws = wb["BoQ"]
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.font.bold is True

    def test_freeze_panes(self, tmp_path: Path):
        doc = _make_procurement_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out)
        wb = _load_workbook(out)
        ws = wb["BoQ"]
        assert ws.freeze_panes == "A2"

    def test_number_format_on_currency(self, tmp_path: Path):
        doc = _make_procurement_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out)
        wb = _load_workbook(out)
        ws = wb["BoQ"]
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value > 100:
                    assert cell.number_format in ('#,##0.00', '#,##0.###', 'General')
                    break


# ── Edge Cases ──────────────────────────────────────────────────────────

class TestEdgeCases:
    def test_empty_document(self, tmp_path: Path):
        doc = GAEBDocument()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out)
        assert out.exists()

    def test_invalid_mode_raises(self, tmp_path: Path):
        doc = _make_procurement_doc()
        out = tmp_path / "test.xlsx"
        with pytest.raises(ValueError, match="Invalid mode"):
            to_excel(doc, out, mode="invalid")

    def test_multi_lot(self, tmp_path: Path):
        items1 = [
            Item(oz="01.0010", short_text="Item A", total_price=Decimal("100")),
        ]
        items2 = [
            Item(oz="01.0010", short_text="Item B", total_price=Decimal("200")),
        ]
        lot1 = Lot(rno="1", label="Lot 1",
                    body=BoQBody(categories=[BoQCtgy(rno="01", label="A", items=items1)]))
        lot2 = Lot(rno="2", label="Lot 2",
                    body=BoQBody(categories=[BoQCtgy(rno="01", label="B", items=items2)]))
        boq = BoQ(lots=[lot1, lot2])
        doc = GAEBDocument(award=AwardInfo(boq=boq))
        out = tmp_path / "test.xlsx"
        to_excel(doc, out)
        wb = _load_workbook(out)
        ws = wb["BoQ"]
        all_vals = [row[0] for row in ws.iter_rows(values_only=True)]
        assert any("Lot 1" in str(v) for v in all_vals if v)
        assert any("Lot 2" in str(v) for v in all_vals if v)

    def test_nested_subcategories(self, tmp_path: Path):
        sub = BoQCtgy(rno="01.01", label="Sub", items=[
            Item(oz="01.01.0010", short_text="Nested item", total_price=Decimal("50")),
        ])
        ctgy = BoQCtgy(rno="01", label="Main", subcategories=[sub])
        lot = Lot(rno="1", label="L", body=BoQBody(categories=[ctgy]))
        boq = BoQ(lots=[lot])
        doc = GAEBDocument(award=AwardInfo(boq=boq))
        out = tmp_path / "test.xlsx"
        to_excel(doc, out)
        wb = _load_workbook(out)
        ws = wb["BoQ"]
        all_vals = [row[0] for row in ws.iter_rows(values_only=True)]
        assert any("Sub" in str(v) for v in all_vals if v)

    def test_alternative_item_type(self, tmp_path: Path):
        items = [
            Item(oz="01.0010", short_text="Normal", item_type=ItemType.NORMAL),
            Item(oz="01.0020", short_text="Alt", item_type=ItemType.ALTERNATIVE),
        ]
        ctgy = BoQCtgy(rno="01", label="A", items=items)
        lot = Lot(rno="1", body=BoQBody(categories=[ctgy]))
        doc = GAEBDocument(award=AwardInfo(boq=BoQ(lots=[lot])))
        out = tmp_path / "test.xlsx"
        to_excel(doc, out)
        wb = _load_workbook(out)
        ws = wb["BoQ"]
        all_vals = []
        for row in ws.iter_rows(values_only=True):
            all_vals.extend(row)
        assert "Alternative" in all_vals

    def test_lazy_import(self):
        from pygaeb import to_excel as te
        assert te is to_excel


# ── Path handling ───────────────────────────────────────────────────────

class TestPathHandling:
    def test_string_path(self, tmp_path: Path):
        doc = _make_procurement_doc()
        out = str(tmp_path / "test.xlsx")
        to_excel(doc, out)
        assert Path(out).exists()

    def test_pathlib_path(self, tmp_path: Path):
        doc = _make_procurement_doc()
        out = tmp_path / "test.xlsx"
        to_excel(doc, out)
        assert out.exists()
