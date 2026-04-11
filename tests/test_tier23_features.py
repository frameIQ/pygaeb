"""Tests for Tier 2 + Tier 3 feature batch.

Covers:
  T2.1 Validation suppression
  T2.2 Document quality score
  T2.3 Diff HTML export
  T2.4 Diff Excel export
  T2.5 GAEB 90 parser
  T3.1 Unit normalization
  T3.2 ÖNORM B 2063 detection
  T3.3 Async parser wrapper
  T3.4 Excel formulas mode
  T3.5 Webhook/event system
  T3.6 PDF export
  T3.7 Database export
  T3.8 XRechnung export
"""

from __future__ import annotations

import asyncio
from decimal import Decimal
from pathlib import Path
from textwrap import dedent

import pytest

from pygaeb import (
    EventType,
    ExchangePhase,
    GAEBParser,
    ItemType,
    SourceVersion,
    aparse,
    clear_subscribers,
    diff_to_excel,
    diff_to_html,
    normalize_unit,
    normalize_units_in_doc,
    on_event,
    quality_score,
    to_excel,
)
from pygaeb.exceptions import GAEBParseError
from pygaeb.models.boq import BoQ, BoQBody, BoQCtgy, Lot
from pygaeb.models.document import AwardInfo, GAEBDocument, GAEBInfo
from pygaeb.models.item import Item

# ── Helpers ──────────────────────────────────────────────────────────


def _make_doc(phase: ExchangePhase = ExchangePhase.X86) -> GAEBDocument:
    items = [
        Item(
            oz="01.0010", short_text="Mauerwerk KS",
            qty=Decimal("100"), unit="m2",
            unit_price=Decimal("45.50"),
            total_price=Decimal("4550.00"),
            item_type=ItemType.NORMAL,
        ),
        Item(
            oz="01.0020", short_text="Beton",
            qty=Decimal("50"), unit="m3",
            unit_price=Decimal("180.00"),
            total_price=Decimal("9000.00"),
            item_type=ItemType.NORMAL,
        ),
    ]
    ctgy = BoQCtgy(rno="01", label="Rohbau", items=items)
    return GAEBDocument(
        source_version=SourceVersion.DA_XML_33,
        exchange_phase=phase,
        gaeb_info=GAEBInfo(version="3.3"),
        award=AwardInfo(
            project_no="P-TEST", project_name="Testprojekt",
            currency="EUR",
            boq=BoQ(lots=[Lot(rno="1", label="Lot", body=BoQBody(categories=[ctgy]))]),
        ),
    )


_SAMPLE_X86 = dedent("""\
    <?xml version="1.0" encoding="utf-8"?>
    <GAEB xmlns="http://www.gaeb.de/GAEB_DA_XML/DA86/3.3">
      <GAEBInfo><Version>3.3</Version></GAEBInfo>
      <Award>
        <AwardInfo><Prj>P1</Prj><PrjName>Test</PrjName><Cur>EUR</Cur></AwardInfo>
        <BoQ><BoQBody><BoQCtgy RNoPart="01"><LblTx>Rohbau</LblTx>
          <Itemlist>
            <Item RNoPart="0010">
              <ShortText>Mauerwerk</ShortText>
              <Qty>100</Qty><QU>m2</QU>
              <UP>45.50</UP><IT>4550.00</IT>
            </Item>
          </Itemlist>
        </BoQCtgy></BoQBody></BoQ>
      </Award>
    </GAEB>
""")


# ═══════════════════════════════════════════════════════════════════════
# T2.1: Validation Warning Suppression
# ═══════════════════════════════════════════════════════════════════════


class TestValidationSuppression:
    def test_suppress_drops_matching_warnings(self, tmp_path: Path) -> None:
        f = tmp_path / "test.X86"
        f.write_text(_SAMPLE_X86)
        # XSD info message normally appears
        doc1 = GAEBParser.parse(str(f))
        xsd_msgs1 = [r for r in doc1.validation_results if "XSD" in r.message]
        assert len(xsd_msgs1) >= 1

        doc2 = GAEBParser.parse(str(f), suppress=["XSD"])
        xsd_msgs2 = [r for r in doc2.validation_results if "XSD" in r.message]
        assert len(xsd_msgs2) == 0

    def test_suppress_does_not_affect_other_warnings(self, tmp_path: Path) -> None:
        f = tmp_path / "test.X86"
        f.write_text(_SAMPLE_X86)
        doc = GAEBParser.parse(str(f), suppress=["nonexistent_pattern"])
        # Original warnings should still be there
        assert len(doc.validation_results) >= 1


# ═══════════════════════════════════════════════════════════════════════
# T2.2: Document Quality Score
# ═══════════════════════════════════════════════════════════════════════


class TestQualityScore:
    def test_score_returns_metrics(self) -> None:
        doc = _make_doc()
        score = quality_score(doc)
        assert 0 <= score.overall <= 100
        assert 0 <= score.completeness <= 100
        assert 0 <= score.precision <= 100
        assert 0 <= score.structure <= 100

    def test_complete_doc_high_score(self) -> None:
        doc = _make_doc()
        score = quality_score(doc)
        assert score.completeness == 100  # All items have text + qty + price

    def test_empty_doc_handled(self) -> None:
        doc = GAEBDocument(
            source_version=SourceVersion.DA_XML_33,
            exchange_phase=ExchangePhase.X86,
            award=AwardInfo(boq=BoQ()),
        )
        score = quality_score(doc)
        assert score.overall >= 0


# ═══════════════════════════════════════════════════════════════════════
# T2.3 + T2.4: Diff HTML/Excel Export
# ═══════════════════════════════════════════════════════════════════════


class TestDiffExports:
    def _make_diff(self) -> object:
        from pygaeb import BoQDiff

        doc_a = _make_doc()
        doc_b = _make_doc()
        # Modify doc_b to create differences
        items = list(doc_b.iter_items())
        items[0].unit_price = Decimal("50.00")
        items[0].total_price = Decimal("5000.00")
        return BoQDiff.compare(doc_a, doc_b)

    def test_diff_html_export(self, tmp_path: Path) -> None:
        result = self._make_diff()
        out = tmp_path / "diff.html"
        diff_to_html(result, out)  # type: ignore[arg-type]
        assert out.exists()
        content = out.read_text()
        assert "<html" in content.lower()
        assert "diff report" in content.lower()
        assert "01.0010" in content

    def test_diff_html_no_changes(self, tmp_path: Path) -> None:
        from pygaeb import BoQDiff

        doc = _make_doc()
        result = BoQDiff.compare(doc, _make_doc())
        out = tmp_path / "nochanges.html"
        diff_to_html(result, out)
        content = out.read_text()
        assert "No item-level changes" in content or "Total changes" in content

    def test_diff_excel_export(self, tmp_path: Path) -> None:
        result = self._make_diff()
        out = tmp_path / "diff.xlsx"
        diff_to_excel(result, out)  # type: ignore[arg-type]
        assert out.exists()

        from openpyxl import load_workbook
        wb = load_workbook(out)
        assert "Summary" in wb.sheetnames
        assert "Items" in wb.sheetnames
        assert "Structure" in wb.sheetnames


# ═══════════════════════════════════════════════════════════════════════
# T2.5: GAEB 90 Parser
# ═══════════════════════════════════════════════════════════════════════


class TestGAEB90Parser:
    def test_parse_minimal_gaeb90(self, tmp_path: Path) -> None:
        # Minimal fixed-width GAEB 90 — line types 00/01/02/03/05/06/07/99
        # Each line is 80 chars; type code is first 2 chars.
        lines = [
            "00" + "PRJ-001             ".ljust(20) + "Testprojekt".ljust(60),
            "01" + "01      ".ljust(8) + "Rohbau".ljust(72),
            "02" + "01.0010       ".ljust(14) + " " * 64,
            "03" + "Mauerwerk KS".ljust(78),
            "05" + "100,000       ".ljust(14) + "m2        ".ljust(10) + " " * 54,
            "06" + "45,50         ".ljust(14) + " " * 64,
            "07" + "4550,00       ".ljust(14) + " " * 64,
            "99" + " " * 78,
        ]
        content = "\n".join(lines)
        f = tmp_path / "legacy.P83"
        f.write_text(content, encoding="utf-8")

        doc = GAEBParser.parse(str(f))
        assert doc.source_version == SourceVersion.GAEB_90
        items = list(doc.iter_items())
        assert len(items) >= 1
        assert items[0].oz == "01.0010"
        assert items[0].short_text == "Mauerwerk KS"
        assert items[0].qty == Decimal("100.000")
        assert items[0].unit == "m2"
        assert items[0].unit_price == Decimal("45.50")

    def test_parse_gaeb90_emits_warning(self, tmp_path: Path) -> None:
        lines = [
            "00" + "P".ljust(78),
            "99" + " " * 78,
        ]
        f = tmp_path / "legacy.P83"
        f.write_text("\n".join(lines), encoding="utf-8")
        doc = GAEBParser.parse(str(f))
        # Should have a warning about minimal parser
        assert any(
            "minimal" in r.message.lower() for r in doc.validation_results
        )


# ═══════════════════════════════════════════════════════════════════════
# T3.1: Unit Normalization
# ═══════════════════════════════════════════════════════════════════════


class TestUnitNormalization:
    def test_normalize_square_meters(self) -> None:
        assert normalize_unit("m2") == "m²"
        assert normalize_unit("qm") == "m²"
        assert normalize_unit("m^2") == "m²"
        assert normalize_unit("m²") == "m²"

    def test_normalize_cubic_meters(self) -> None:
        assert normalize_unit("m3") == "m³"
        assert normalize_unit("cbm") == "m³"
        assert normalize_unit("m^3") == "m³"

    def test_normalize_pieces(self) -> None:
        assert normalize_unit("Stk") == "Stk"
        assert normalize_unit("Stck") == "Stk"
        assert normalize_unit("st") == "Stk"
        assert normalize_unit("pcs") == "Stk"

    def test_normalize_lump_sum(self) -> None:
        assert normalize_unit("psch") == "psch"
        assert normalize_unit("Pauschal") == "psch"

    def test_normalize_unknown_unchanged(self) -> None:
        assert normalize_unit("unknown_unit") == "unknown_unit"

    def test_normalize_none_returns_none(self) -> None:
        assert normalize_unit(None) is None

    def test_normalize_empty_string(self) -> None:
        assert normalize_unit("") == ""

    def test_normalize_units_in_doc(self) -> None:
        doc = _make_doc()
        # Items already have m2, m3 → will be changed to m², m³
        changed = normalize_units_in_doc(doc)
        assert changed == 2
        items = list(doc.iter_items())
        assert items[0].unit == "m²"
        assert items[1].unit == "m³"


# ═══════════════════════════════════════════════════════════════════════
# T3.2: ÖNORM B 2063 Detection
# ═══════════════════════════════════════════════════════════════════════


class TestONORMDetection:
    def test_onorm_extension_raises_graceful_error(self, tmp_path: Path) -> None:
        f = tmp_path / "test.onlv"
        f.write_text("ONORM B 2063 file content")
        with pytest.raises(GAEBParseError, match="ÖNORM"):
            GAEBParser.parse(str(f))

    def test_onorm_header_detected(self, tmp_path: Path) -> None:
        f = tmp_path / "test.xml"
        f.write_text("ONORM B 2063 content here\nmore content")
        with pytest.raises(GAEBParseError, match="ÖNORM"):
            GAEBParser.parse(str(f))


# ═══════════════════════════════════════════════════════════════════════
# T3.3: Async Parser Wrapper
# ═══════════════════════════════════════════════════════════════════════


class TestAsyncParser:
    @pytest.mark.asyncio
    async def test_aparse_works(self, tmp_path: Path) -> None:
        f = tmp_path / "test.X86"
        f.write_text(_SAMPLE_X86)
        doc = await aparse(str(f))
        assert doc.exchange_phase == ExchangePhase.X86
        assert len(list(doc.iter_items())) == 1

    @pytest.mark.asyncio
    async def test_aparse_concurrent(self, tmp_path: Path) -> None:
        f = tmp_path / "test.X86"
        f.write_text(_SAMPLE_X86)
        # Parse 3 docs concurrently
        results = await asyncio.gather(
            aparse(str(f)), aparse(str(f)), aparse(str(f)),
        )
        assert len(results) == 3
        assert all(d.exchange_phase == ExchangePhase.X86 for d in results)


# ═══════════════════════════════════════════════════════════════════════
# T3.4: Excel Formulas Mode
# ═══════════════════════════════════════════════════════════════════════


class TestExcelFormulas:
    def test_formulas_mode_writes_excel_formulas(self, tmp_path: Path) -> None:
        doc = _make_doc()
        out = tmp_path / "out.xlsx"
        to_excel(doc, out, formulas=True)

        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb.active
        # Find a cell containing "=" formula in the data area
        formulas_found = False
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas_found = True
                    break
            if formulas_found:
                break
        assert formulas_found, "Expected at least one =formula cell"

    def test_no_formulas_mode_writes_static_values(self, tmp_path: Path) -> None:
        doc = _make_doc()
        out = tmp_path / "out_static.xlsx"
        to_excel(doc, out, formulas=False)

        from openpyxl import load_workbook
        wb = load_workbook(out)
        ws = wb.active
        for row in ws.iter_rows():
            for cell in row:
                assert not (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                ), "Static mode should not have formulas"


# ═══════════════════════════════════════════════════════════════════════
# T3.5: Webhook/Event System
# ═══════════════════════════════════════════════════════════════════════


class TestEventSystem:
    def setup_method(self) -> None:
        clear_subscribers()

    def teardown_method(self) -> None:
        clear_subscribers()

    def test_parse_completed_event(self, tmp_path: Path) -> None:
        events_received: list[dict[str, object]] = []

        def callback(payload: dict[str, object]) -> None:
            events_received.append(payload)

        on_event(EventType.PARSE_COMPLETED, callback)

        f = tmp_path / "test.X86"
        f.write_text(_SAMPLE_X86)
        GAEBParser.parse(str(f))

        assert len(events_received) == 1
        assert events_received[0]["item_count"] == 1
        assert events_received[0]["phase"] == "X86"

    def test_callback_failure_does_not_break_parse(self, tmp_path: Path) -> None:
        def bad_callback(payload: dict[str, object]) -> None:
            raise ValueError("Boom")

        on_event(EventType.PARSE_COMPLETED, bad_callback)
        f = tmp_path / "test.X86"
        f.write_text(_SAMPLE_X86)
        # Should not raise despite the failing callback
        doc = GAEBParser.parse(str(f))
        assert doc is not None

    def test_multiple_subscribers(self, tmp_path: Path) -> None:
        counter = {"count": 0}

        def cb1(p: dict[str, object]) -> None:
            counter["count"] += 1

        def cb2(p: dict[str, object]) -> None:
            counter["count"] += 10

        on_event(EventType.PARSE_COMPLETED, cb1)
        on_event(EventType.PARSE_COMPLETED, cb2)

        f = tmp_path / "test.X86"
        f.write_text(_SAMPLE_X86)
        GAEBParser.parse(str(f))

        assert counter["count"] == 11


# ═══════════════════════════════════════════════════════════════════════
# T3.6: PDF Export
# ═══════════════════════════════════════════════════════════════════════


class TestPDFExport:
    def test_pdf_import_error_when_missing(self, tmp_path: Path) -> None:
        """PDF export should raise ImportError if reportlab is not installed."""
        try:
            import reportlab  # noqa: F401
            pytest.skip("reportlab is installed; cannot test ImportError")
        except ImportError:
            pass

        from pygaeb.convert.to_pdf import to_pdf

        doc = _make_doc()
        with pytest.raises(ImportError, match="reportlab"):
            to_pdf(doc, tmp_path / "out.pdf")

    def test_pdf_export_with_reportlab(self, tmp_path: Path) -> None:
        """If reportlab is installed, PDF generation should work."""
        try:
            import reportlab  # noqa: F401
        except ImportError:
            pytest.skip("reportlab not installed")

        from pygaeb.convert.to_pdf import to_pdf

        doc = _make_doc()
        out = tmp_path / "out.pdf"
        to_pdf(doc, out)
        assert out.exists()
        # PDF files start with %PDF-
        assert out.read_bytes().startswith(b"%PDF-")


# ═══════════════════════════════════════════════════════════════════════
# T3.7: Database Export
# ═══════════════════════════════════════════════════════════════════════


class TestDatabaseExport:
    def test_database_import_error_when_missing(self, tmp_path: Path) -> None:
        try:
            import sqlalchemy  # noqa: F401
            pytest.skip("sqlalchemy is installed; cannot test ImportError")
        except ImportError:
            pass

        from pygaeb.convert.to_database import to_database

        doc = _make_doc()
        with pytest.raises(ImportError, match="sqlalchemy"):
            to_database(doc, None)  # type: ignore[arg-type]

    def test_database_export_to_sqlite(self, tmp_path: Path) -> None:
        try:
            from sqlalchemy import create_engine, select
        except ImportError:
            pytest.skip("sqlalchemy not installed")

        from pygaeb.convert.to_database import (
            _get_metadata,
            create_schema,
            to_database,
        )

        db_path = tmp_path / "test.db"
        engine = create_engine(f"sqlite:///{db_path}")
        create_schema(engine)

        doc = _make_doc()
        doc_id = to_database(doc, engine)
        assert doc_id > 0

        # Verify rows were inserted
        _, documents, _, items = _get_metadata()
        with engine.connect() as conn:
            doc_row = conn.execute(
                select(documents).where(documents.c.id == doc_id)
            ).first()
            assert doc_row is not None
            assert doc_row.project_no == "P-TEST"

            item_rows = conn.execute(
                select(items).where(items.c.document_id == doc_id)
            ).all()
            assert len(item_rows) == 2


# ═══════════════════════════════════════════════════════════════════════
# T3.8: XRechnung Export
# ═══════════════════════════════════════════════════════════════════════


class TestXRechnungExport:
    def test_xrechnung_requires_x89(self, tmp_path: Path) -> None:
        from pygaeb.convert.to_xrechnung import to_xrechnung

        doc = _make_doc(phase=ExchangePhase.X86)  # not X89
        with pytest.raises(ValueError, match="X89"):
            to_xrechnung(doc, tmp_path / "out.xml")

    def test_xrechnung_x89_export(self, tmp_path: Path) -> None:
        from pygaeb.convert.to_xrechnung import to_xrechnung

        doc = _make_doc(phase=ExchangePhase.X89)
        out = tmp_path / "invoice_xrechnung.xml"
        to_xrechnung(
            doc, out,
            invoice_number="INV-001",
            seller_name="Acme Bau GmbH",
            buyer_name="Stadtverwaltung Musterstadt",
            leitweg_id="04011000-1234567890-12",
        )
        assert out.exists()

        content = out.read_text()
        assert "Invoice" in content
        assert "INV-001" in content
        assert "Acme Bau GmbH" in content
        assert "04011000-1234567890-12" in content
        assert "VAT" in content

    def test_xrechnung_includes_items(self, tmp_path: Path) -> None:
        from pygaeb.convert.to_xrechnung import to_xrechnung

        doc = _make_doc(phase=ExchangePhase.X89)
        out = tmp_path / "invoice.xml"
        to_xrechnung(doc, out)

        content = out.read_text()
        assert "InvoiceLine" in content
        assert "Mauerwerk" in content
